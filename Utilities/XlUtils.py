
import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file,sheetName):
    wb=openpyxl.load_workbook(file)
    sheet=wb[sheetName]
    return sheet.max_row

def getColumnCount(file,sheetName):
    wb=openpyxl.load_workbook(file)
    sheet=wb[sheetName]
    return sheet.max_column

def getReadData(file,sheetName,row,column):
    wb=openpyxl.load_workbook(file)
    sheet=wb[sheetName]
    return sheet.cell(row=row,column=column).value

def getWriteData(file,sheetName,row,column,data):
    wb=openpyxl.load_workbook(file)
    sheet=wb[sheetName]
    sheet.cell(row=row,column=column).value=data
    wb.save(file)

def getFillColor(file,sheetName,row,column,color):
    wb=openpyxl.load_workbook(file)
    sheet=wb[sheetName]
    sheet.cell(row=row,column=column) .fill=PatternFill(start_color=color,end_color=color,fill_type='solid')
    wb.save(file)

def getFillGreenColor(file,sheetName,rownum,columnnum):
    wb=openpyxl.load_workbook(file)
    sheet=wb[sheetName]
    sheet.cell(row=rownum,column=columnnum).fill=PatternFill(start_color='60b212',end_color='60b212',fill_type='solid')
    wb.save(file)
    return
def getFillRedColor(file,sheetName,rownum,columnnum):
    wb=openpyxl.load_workbook(file)
    sheet=wb[sheetName]
    sheet.cell(row=rownum,column=columnnum).fill=PatternFill(start_color='FF0000',end_color='FF0000',fill_type='solid')
    wb.save(file)
    return




